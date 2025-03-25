import os
import json
import time
from urllib.parse import quote
from bs4 import BeautifulSoup
import requests
from difflib import get_close_matches
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SEARCH_CACHE = "data/rs3_search_cache.json"
SEARCH_META = "data/rs3_search_cache_meta.txt"
CACHE_EXPIRY = 4 * 60 * 60  # 4 hours

HEADERS = {
    "User-Agent": "RS3-Market-Watcher/1.0 (by S3venScars)"
}

WIKI_EXCHANGE_URL = "https://runescape.wiki/w/"


def format_url_from_name(item_name: str) -> str:
    slug = quote(item_name.replace(" ", "_"))
    return f"{WIKI_EXCHANGE_URL}Exchange:{slug}"


def parse_exchange_page(url: str) -> dict:
    response = requests.get(url, headers=HEADERS)
    if response.status_code != 200:
        raise Exception(f"Failed to fetch: {url}")

    soup = BeautifulSoup(response.text, "lxml")
    def get_stat_by_id(stat_id):
        tag = soup.find(id=stat_id)
        if tag:
            return tag.text.strip().replace(",", "")
        return None

    return {
        "id": int(get_stat_by_id("exchange-itemid") or 0),
        "ge_price": int(get_stat_by_id("GEPrice") or 0),
        "high_alch": int(get_stat_by_id("exchange-highalch") or 0),
        "low_alch": int(get_stat_by_id("exchange-lowalch") or 0),
        "name": soup.find("h1").text.strip()
    }


def is_cache_expired() -> bool:
    if not os.path.exists(SEARCH_META):
        return True
    try:
        with open(SEARCH_META, "r") as f:
            last_ts = float(f.read().strip())
        return (time.time() - last_ts) > CACHE_EXPIRY
    except Exception:
        return True


def update_cache_timestamp():
    with open(SEARCH_META, "w") as f:
        f.write(str(time.time()))


def load_search_cache():
    if not os.path.exists(SEARCH_CACHE):
        return {}
    with open(SEARCH_CACHE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_search_cache(cache: dict):
    os.makedirs("data", exist_ok=True)
    with open(SEARCH_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2)


def fuzzy_search_items(term: str) -> list:
    cache = load_search_cache()
    keys = cache.keys()
    matches = [key for key in keys if term.lower() in key.lower()]
    return [cache[m] for m in matches]


def search_item(item_name: str) -> dict:
    cache = load_search_cache()

    if is_cache_expired():
        cache.clear()
        update_cache_timestamp()

    key = item_name.lower()
    if key in cache:
        return cache[key]

    url = format_url_from_name(item_name)
    data = parse_exchange_page(url)
    data["url"] = url
    data["last_updated"] = datetime.utcnow().isoformat()
    cache[key] = data
    save_search_cache(cache)
    return data


def clear_search_cache():
    if os.path.exists(SEARCH_CACHE):
        os.remove(SEARCH_CACHE)
    if os.path.exists(SEARCH_META):
        os.remove(SEARCH_META)


def export_cache_to_excel(filepath="watch_cache.xlsx"):
    cache = load_search_cache()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RS3 Watchlist Cache"

    headers = ["Name", "ID", "GE Price", "High Alch", "Alch Profit", "URL", "Last Updated"]
    ws.append(headers)

    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 20

    for data in cache.values():
        ge = data.get("ge_price", 0)
        ha = data.get("high_alch", 0)
        row = [
            data["name"],
            data["id"],
            ge,
            ha,
            ha - ge,
            data.get("url", ""),
            data.get("last_updated", "")
        ]
        ws.append(row)

    wb.save(filepath)
